import csv
import openpyxl
import uuid
import logging
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.contrib import messages
from django.http import HttpResponse
from django.views import View
from django.db import transaction
from django.contrib.auth import get_user_model
from django.utils.crypto import get_random_string
from django.db.models import Q

from Exactus.employee.forms import get_employee_form_for_country, EmployeeUploadForm


# Models
from Exactus.company.models import Company
from Exactus.employee.models import Employee
from Exactus.country.models import Country

# Forms
from Exactus.employee.forms import (
    get_employee_form_for_country,
    EmployeeUploadForm,
    EmployeeAccessForm,
    CompensationForm,
    EmployeeAccessForm,
)

# Utils & Permissions
from Exactus.accounts.utils.decorators import role_required

# Service Import (Robust handling)
try:
    from Exactus.accounts.services.onboarding import OnboardingService
except ImportError:
    # Log error if service is missing so it doesn't crash the app on load
    logger = logging.getLogger(__name__)
    logger.error("Could not import OnboardingService. Check file Exactus/accounts/services/onboarding.py")
    OnboardingService = None

User = get_user_model()

# ──────────────────────────────────────────────────────────────────────────────
# SECURITY HELPER
# ──────────────────────────────────────────────────────────────────────────────

def validate_company_access(user, company):
    """
    Verifies if the current user is allowed to access the specific company.
    Logic:
    1. Global Roles (EXEC, ADMIN, COMPLIANCE) -> ALLOWED.
    2. Restricted Roles -> Must be linked via 'user.contexts'.
    """
    global_roles = ["EXEC", "ADMIN", "COMPLIANCE"]
    user_role = getattr(user, 'role', '').upper()

    # 1. Superusers and Global Roles always pass
    if user.is_superuser or user_role in global_roles:
        return True

    # 2. Check for direct assignment
    if user.contexts.filter(company=company).exists():
        return True

    return False

# ──────────────────────────────────────────────────────────────────────────────
# EMPLOYEE CRUD
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def employee_list(request, country_slug, company_id):
    """
    List employees for a specific company.
    """
    country = get_object_or_404(Country, slug=country_slug)
    company = get_object_or_404(Company, pk=company_id)

    # Security Check
    if not validate_company_access(request.user, company):
        messages.error(request, "Access Denied: You are not assigned to this company.")
        return redirect("dashboard")

    employees = Employee.objects.filter(company=company).order_by("employee_number")

    return render(request, "employee/index.html", {
        "company": company,
        "employees": employees,
        "country": country,
        "country_slug": country_slug,
    })

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def employee_create(request, country_slug, company_id):
    """
    Create a new employee record.
    """
    country = get_object_or_404(Country, slug=country_slug)
    company = get_object_or_404(Company, pk=company_id)

    if not validate_company_access(request.user, company):
        messages.error(request, "Access Denied.")
        return redirect("dashboard")

    FormClass = get_employee_form_for_country(country)

    if request.method == "POST":
        form = FormClass(request.POST)
        if form.is_valid():
            employee = form.save(commit=False)
            employee.company = company
            employee.save()
            messages.success(request, f"Employee '{employee.employee_name}' added successfully.")
            return redirect('employee:employee', country_slug=country_slug, company_id=company.company_id)
    else:
        form = FormClass()

    return render(request, "employee/form.html", {
        "form": form, "company": company, "country": country, "country_slug": country_slug
    })



import logging

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordResetForm
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render

from Exactus.accounts.utils.decorators import role_required
from Exactus.company.models import Company
from Exactus.compensation.models import CompensationComponent
from Exactus.country.models import Country
from Exactus.employee.models import Employee
from Exactus.employee.forms import get_employee_form_for_country
from Exactus.employee.forms.access_form import EmployeeAccessForm

logger = logging.getLogger(__name__)


# -------------------------------------------------------------------
# Access helpers
# -------------------------------------------------------------------
def validate_company_access(user, company: Company) -> bool:
    """
    Verifica se o user tem acesso à company.
    - Superuser e roles globais: permitido
    - Caso contrário: precisa estar ligado via user.contexts (se existir)
    """
    global_roles = {"EXEC", "ADMIN", "COMPLIANCE"}
    user_role = (getattr(user, "role", "") or "").upper()

    if getattr(user, "is_superuser", False) or user_role in global_roles:
        return True

    # Se o teu projeto usa user.contexts para vínculos:
    if hasattr(user, "contexts"):
        try:
            return user.contexts.filter(company=company).exists()
        except Exception:
            pass

    return False


# -------------------------------------------------------------------
# Onboarding helpers
# -------------------------------------------------------------------
def _pick_default_employee_role(UserModel):
    """
    Escolhe um role default para um colaborador, se o teu User model tiver isso.
    """
    candidates = ("EMPLOYEE", "Employee", "CLIENT_EMPLOYEE", "SPECIALIST")

    if hasattr(UserModel, "CLIENT_ROLES"):
        allowed = set(getattr(UserModel, "CLIENT_ROLES") or [])
        for c in candidates:
            if c in allowed:
                return c

    if hasattr(UserModel, "ROLE_CHOICES"):
        role_codes = [code for code, _ in getattr(UserModel, "ROLE_CHOICES") or []]
        for c in candidates:
            if c in role_codes:
                return c
        if role_codes:
            return role_codes[0]

    return None


def _send_welcome_or_setup_email(user, request) -> bool:
    """
    Envia um link seguro de definição de password.
    Fallback: PasswordResetForm (funciona como "set password").
    """
    try:
        # Fallback robusto (sem depender do teu OnboardingService)
        prf = PasswordResetForm({"email": user.email})
        if prf.is_valid():
            prf.save(
                request=request,
                use_https=request.is_secure(),
                from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
            )
            return True
    except Exception:
        logger.exception("Failed to send setup email via PasswordResetForm")

    return False


# -------------------------------------------------------------------
# View
# -------------------------------------------------------------------
@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def employee_edit(request, country_slug, company_id, employee_id):
    """
    Edita um Employee e (opcionalmente) cria a conta do portal para o colaborador.
    - POST normal: salva employee e, se existir, salva access_form (role/client_group/is_active)
    - POST com create_user_account=true: salva employee, cria user se não existir, envia email, volta para a página.
    """
    User = get_user_model()

    country = get_object_or_404(Country, slug=country_slug)
    company = get_object_or_404(Company, pk=company_id)

    if not validate_company_access(request.user, company):
        messages.error(request, "Access Denied.")
        return redirect("dashboard")

    employee = get_object_or_404(Employee, pk=employee_id, company=company)
    FormClass = get_employee_form_for_country(country)

    active_components = (
        CompensationComponent.objects.filter(employee=employee, processed=False)
        .exclude(pd_code__pdcode_status="Hidden")
        .select_related("pd_code", "element")
    )

    linked_user = User.objects.filter(email__iexact=employee.email).first() if employee.email else None

    if request.method == "POST":
        create_account_clicked = request.POST.get("create_user_account") == "true"

        form = FormClass(request.POST, instance=employee)
        access_form = EmployeeAccessForm(request.POST, instance=linked_user) if linked_user else None

        if not form.is_valid():
            messages.error(request, "Please fix the highlighted errors and try again.")
            return render(
                request,
                "employee/form.html",
                {
                    "form": form,
                    "access_form": access_form,
                    "linked_user": linked_user,
                    "employee": employee,
                    "company": company,
                    "country": country,
                    "country_slug": country_slug,
                    "active_components": active_components,
                },
            )

        with transaction.atomic():
            employee = form.save()

            # recalc depois do save (email pode ter sido editado)
            email = (employee.email or "").strip()
            linked_user = User.objects.filter(email__iexact=email).first() if email else None

            # 1) criar conta
            if create_account_clicked:
                if not email:
                    messages.error(request, "Employee email is required to create an account.")
                    return redirect(
                        "employee:employee_edit",
                        country_slug=country_slug,
                        company_id=company.pk,
                        employee_id=employee.pk,
                    )

                if linked_user:
                    messages.info(request, "A user account already exists for this email.")
                    return redirect(
                        "employee:employee_edit",
                        country_slug=country_slug,
                        company_id=company.pk,
                        employee_id=employee.pk,
                    )

                # username seguro + único
                base = email.split("@")[0][:30]
                username = base
                i = 1
                while User.objects.filter(username=username).exists():
                    i += 1
                    username = f"{base[:25]}{i}"

                # criar user (compatível com custom user model)
                if hasattr(User.objects, "create_user"):
                    user = User.objects.create_user(username=username, email=email)
                else:
                    user = User.objects.create(username=username, email=email)

                if hasattr(user, "is_active"):
                    user.is_active = True
                if hasattr(user, "set_unusable_password"):
                    user.set_unusable_password()

                if hasattr(user, "role") and not getattr(user, "role", None):
                    default_role = _pick_default_employee_role(User)
                    if default_role:
                        user.role = default_role

                user.save()

                sent = _send_welcome_or_setup_email(user, request)
                if sent:
                    messages.success(request, f"Account created and email sent to {email}.")
                else:
                    messages.warning(request, f"Account created for {email}, but email could not be sent.")

                return redirect(
                    "employee:employee_edit",
                    country_slug=country_slug,
                    company_id=company.pk,
                    employee_id=employee.pk,
                )

            # 2) salvar access_form (se existir)
            if access_form:
                if access_form.is_valid():
                    access_form.save()
                else:
                    messages.error(request, "Employee saved, but access settings contain errors.")
                    return render(
                        request,
                        "employee/form.html",
                        {
                            "form": form,
                            "access_form": access_form,
                            "linked_user": linked_user,
                            "employee": employee,
                            "company": company,
                            "country": country,
                            "country_slug": country_slug,
                            "active_components": active_components,
                        },
                    )

        messages.success(request, "Employee updated successfully.")
        return redirect("employee:employee", country_slug=country_slug, company_id=company.company_id)

    # GET
    form = FormClass(instance=employee)
    linked_user = User.objects.filter(email__iexact=employee.email).first() if employee.email else None
    access_form = EmployeeAccessForm(instance=linked_user) if linked_user else None

    return render(
        request,
        "employee/form.html",
        {
            "form": form,
            "access_form": access_form,
            "linked_user": linked_user,
            "employee": employee,
            "company": company,
            "country": country,
            "country_slug": country_slug,
            "active_components": active_components,
        },
    )






# ──────────────────────────────────────────────────────────────────────────────
# MULTI-FORMAT UPLOAD (Class-Based View)
# ──────────────────────────────────────────────────────────────────────────────

class EmployeeUploadView(View):
    """Handle bulk employee uploads via CSV or Excel."""

    @method_decorator(login_required)
    @method_decorator(role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST"))
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, country_slug=None, company_id=None):
        country = get_object_or_404(Country, slug=country_slug) if country_slug else None
        company = get_object_or_404(Company, pk=company_id) if company_id else None

        # Security Check for GET
        if company and not validate_company_access(request.user, company):
            messages.error(request, "Access Denied.")
            return redirect("dashboard")

        return render(request, "employee/upload_form.html", {
            "form": EmployeeUploadForm(),
            "company": company,
            "country": country,
            "country_slug": country_slug
        })

    @transaction.atomic
    def post(self, request, country_slug=None, company_id=None):
        country = get_object_or_404(Country, slug=country_slug) if country_slug else None
        company = get_object_or_404(Company, pk=company_id) if company_id else None

        # Security Check for POST
        if company and not validate_company_access(request.user, company):
            messages.error(request, "Access Denied.")
            return redirect("dashboard")

        form = EmployeeUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            return self.get(request, country_slug, company_id)

        file = request.FILES["file"]
        filename = file.name.lower()
        dry_run = form.cleaned_data.get("dry_run", False)

        try:
            data_rows = []

            # --- CSV Processing ---
            if filename.endswith('.csv'):
                file.seek(0)
                content = file.read().decode('utf-8-sig').splitlines()
                reader = csv.reader(content)
                rows = list(reader)
                if len(rows) < 2:
                    messages.error(request, "The CSV file is empty.")
                    return self.get(request, country_slug, company_id)
                headers = [str(h).strip().lower() for h in rows[0]]
                for row in rows[1:]:
                    if any(row):
                        data_rows.append(dict(zip(headers, row)))

            # --- Excel Processing ---
            elif filename.endswith(('.xlsx', '.xls')):
                file.seek(0)
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    messages.error(request, "The Excel file is empty.")
                    return self.get(request, country_slug, company_id)
                headers = [str(h).strip().lower() for h in rows[0]]
                for row in rows[1:]:
                    if any(row):
                        data_rows.append(dict(zip(headers, row)))
            else:
                messages.error(request, "Unsupported format. Please upload .xlsx or .csv")
                return self.get(request, country_slug, company_id)

            success_count = 0
            error_log = []

            # Cache companies to avoid repeated DB hits if global upload
            company_cache = {c.company_code: c for c in Company.objects.all()}

            for index, data in enumerate(data_rows, start=2):
                try:
                    # Logic: If company_id is provided in URL, force that company.
                    # Otherwise, look for 'company_code' in the file.
                    if company:
                        target_company = company
                    else:
                        c_code = str(data.get('company_code', '')).strip()
                        target_company = company_cache.get(c_code)
                        if not target_company:
                            error_log.append(f"Row {index}: Company code '{c_code}' not found.")
                            continue

                    emp_num = data.get('employee_number')
                    if not emp_num:
                        error_log.append(f"Row {index}: Missing employee_number.")
                        continue

                    emp_code = data.get('employee_code') or emp_num

                    defaults = {
                        'employee_name': data.get('employee_name'),
                        'employee_surname': data.get('employee_surname'),
                        'employee_code': emp_code,
                        'gender': data.get('gender'),
                        'email': data.get('email'),
                        'date_of_birth': data.get('date_of_birth'),
                        'employment_start_date': data.get('employment_start_date'),
                    }

                    # Optional fields mapping
                    if 'ni_number' in data: defaults['tax_info_01'] = data['ni_number']
                    if 'tax_code' in data: defaults['tax_info_03'] = data['tax_code']
                    if 'cpf' in data: defaults['tax_info_01'] = data['cpf']
                    if 'pis' in data: defaults['tax_info_02'] = data['pis']

                    if not dry_run:
                        Employee.objects.update_or_create(
                            company=target_company,
                            employee_number=emp_num,
                            defaults=defaults
                        )
                    success_count += 1
                except Exception as e:
                    error_log.append(f"Row {index}: {str(e)}")

            request.session["upload_result"] = {
                "success_count": success_count,
                "errors": error_log,
                "total": len(data_rows),
                "dry_run": dry_run
            }

            if company_id:
                return redirect("employee:employee_upload_result", country_slug=country_slug, company_id=company_id)
            return redirect("employee:global_upload_result", country_slug=country_slug)

        except Exception as e:
            messages.error(request, f"File processing error: {str(e)}")
            return self.get(request, country_slug, company_id)

# ──────────────────────────────────────────────────────────────────────────────
# RESULT VIEWS
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def employee_upload_result_view(request, country_slug=None, company_id=None):
    """Display upload outcome."""
    result = request.session.pop("upload_result", None)
    if not result:
        messages.warning(request, "No upload results found.")
        return redirect("dashboard")

    country = get_object_or_404(Country, slug=country_slug) if country_slug else None
    company = get_object_or_404(Company, pk=company_id) if company_id else None

    if company and not validate_company_access(request.user, company):
        messages.error(request, "Access Denied.")
        return redirect("dashboard")

    return render(request, "employee/upload_result.html", {
        "result": result,
        "company": company,
        "country": country,
        "country_slug": country_slug
    })

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def global_upload_result_view(request, country_slug):
    """Alias for global route consistency."""
    return employee_upload_result_view(request, country_slug=country_slug)

# ──────────────────────────────────────────────────────────────────────────────
# TEMPLATE DOWNLOAD
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def download_employees_template(request, country_slug, company_id=None):
    """Generate Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Import"

    headers = [
        'company_code', 'employee_number', 'employee_code', 'employee_name',
        'employee_surname', 'gender', 'date_of_birth', 'email', 'employment_start_date'
    ]

    # Country-specific fields
    if country_slug in ['united-kingdom', 'uk', 'gb']:
        headers += ['ni_number', 'tax_code']
    elif country_slug in ['brazil', 'br']:
        headers += ['cpf', 'pis']

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    if company_id:
        company = get_object_or_404(Company, pk=company_id)
        # Security check before pre-filling
        if validate_company_access(request.user, company):
            ws.cell(row=2, column=1, value=company.company_code)

    gender_dv = DataValidation(type="list", formula1='"Male,Female"', allow_blank=True)
    ws.add_data_validation(gender_dv)
    gender_dv.add("F2:F500")

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename=Employee_Template_{country_slug}.xlsx'
    wb.save(response)
    return response

# ... (imports) ...
from Exactus.employee.models import Compensation
from Exactus.employee.forms import CompensationForm

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def employee_compensation(request, country_slug, company_id, employee_id):
    """
    Manage Compensation History for an Employee.
    """
    country = get_object_or_404(Country, slug=country_slug)
    company = get_object_or_404(Company, pk=company_id)

    if not validate_company_access(request.user, company):
        messages.error(request, "Access Denied.")
        return redirect("dashboard")

    employee = get_object_or_404(Employee, pk=employee_id, company=company)
    compensations = employee.compensations.all() # Ordered by -effective_date via Meta

    if request.method == "POST":
        form = CompensationForm(request.POST)
        if form.is_valid():
            comp_record = form.save(commit=False)
            comp_record.employee = employee
            comp_record.save()
            messages.success(request, "Compensation record added successfully.")
            return redirect("employee:employee_compensation", country_slug=country_slug, company_id=company_id, employee_id=employee_id)
    else:
        form = CompensationForm()

    return render(request, "employee/compensation.html", {
        "employee": employee,
        "company": company,
        "country": country,
        "country_slug": country_slug,
        "compensations": compensations,
        "form": form
    })


def validate_company_access(user, company):
    """
    Verifies if the current user is allowed to access the specific company.
    Logic:
    1. Global Roles (EXEC, ADMIN, COMPLIANCE) -> ALLOWED.
    2. Restricted Roles -> Must be linked via 'user.contexts'.
    """
    global_roles = ["EXEC", "ADMIN", "COMPLIANCE"]
    user_role = getattr(user, "role", "").upper()

    if user.is_superuser or user_role in global_roles:
        return True

    if user.contexts.filter(company=company).exists():
        return True

    return False