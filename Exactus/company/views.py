import csv
import json
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.db.models import ProtectedError
# Models
from Exactus.company.models import Company
from Exactus.country.models import Country

# Forms & Helpers
from Exactus.company.registry import get_company_form_class   
from Exactus.company.utils.csv_importer import import_from_csv
from Exactus.company.forms.utils import get_company_form_for_country
# Added missing import
from Exactus.company.forms import CompanyUploadForm 

# Permissions - FIXED: Use the v3.1 compatible decorator we updated earlier
from Exactus.country.utils.decorators import role_required

# ────────────────────────────────────────────────────────────────
# 🧩 Company Index
# ────────────────────────────────────────────────────────────────

from django.db.models import Count, Q, Exists, OuterRef
from Exactus.payroll.models import Payroll  # Import Payroll model



@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "BILLING", "IMPLEMENTATION",
               "OPERATION", "DIRECTOR", "MANAGER", "SPECIALIST", "FINANCE")
def company(request, country_slug):
    country = get_object_or_404(Country, slug=country_slug)
    
    GLOBAL_ACCESS_ROLES = ['EXEC', 'ADMIN', 'OPERATION', 'COMPLIANCE', 'BILLING', 'IMPLEMENTATION']
    
    # 1. Base Query
    if request.user.is_staff or request.user.role in GLOBAL_ACCESS_ROLES:
        queryset = Company.objects.filter(country=country)
    else:
        queryset = Company.objects.filter(
            country=country,
            authorized_users__user=request.user,
            authorized_users__is_active=True
        ).distinct()

    # 2. Add Annotations (Employees Count & Open Payroll Status)
    # We assume 'open' payrolls are those NOT completed/paid. Adjust status keys as needed.
    OPEN_STATUSES = ['DRAFT', 'OPEN', 'PROCESSING', 'review', 'pending'] 
    
    companies = queryset.annotate(
        # Count all employees linked to this company
        total_employees=Count('employees', distinct=True),
        
        # Check if ANY payroll exists with an 'open' status
        has_open_payroll=Exists(
            Payroll.objects.filter(
                company=OuterRef('pk'),
                status__in=OPEN_STATUSES 
            )
        )
    ).order_by("trade_name")

    return render(request, "company/index.html", {
        "country": country,
        "companies": companies,
        "country_slug": country.slug
    })


# ────────────────────────────────────────────────────────────────
# ➕ Create Company
# ────────────────────────────────────────────────────────────────

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "BILLING", "IMPLEMENTATION", "OPERATION")
def company_create(request, country_slug):
    country = get_object_or_404(Country, slug=country_slug)
    
    # Get the specific class (e.g., UnitedKingdomCompanyForm)
    FormClass = get_company_form_for_country(country)

    if request.method == "POST":
        form = FormClass(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                instance = form.save(commit=False)
                instance.country = country  # Attach country manually
                instance.save()
                messages.success(request, f"Company '{instance.trade_name}' created successfully!")
                return redirect("companies:company", country.slug)
            except Exception as e:
                messages.error(request, f"Error saving company: {str(e)}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = FormClass()

    return render(request, "company/form.html", {
        "form": form,
        "country": country,
        "company": None,
    })


# ────────────────────────────────────────────────────────────────
# ✏️ Edit Company
# ────────────────────────────────────────────────────────────────

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "BILLING", "IMPLEMENTATION", "OPERATION")
def company_edit(request, country_slug, company_id):
    country = get_object_or_404(Country, slug=country_slug)
    company = get_object_or_404(Company, pk=company_id)
    
    # Get the specific class
    FormClass = get_company_form_for_country(country)

    if request.method == "POST":
        form = FormClass(request.POST, request.FILES, instance=company)
        
        if form.is_valid():
            try:
                instance = form.save(commit=False)
                instance.country = country
                instance.save()
                messages.success(request, f"Company '{instance.trade_name}' updated successfully!")
                return redirect("companies:company", country.slug)
            except Exception as e:
                messages.error(request, f"Error updating company: {str(e)}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = FormClass(instance=company)

    return render(request, "company/form.html", {
        "form": form,
        "country": country,
        "company": company,
    })


# ────────────────────────────────────────────────────────────────
# 🗑 Delete Listing
# ────────────────────────────────────────────────────────────────

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models.signals import pre_delete

# 1. Import the specific signal and model causing the block
from Exactus.payroll.signals import prevent_delete_processed_period
from Exactus.payroll.models import PayrollPeriod
from Exactus.country.models import Country
from Exactus.company.models import Company
from Exactus.country.utils.decorators import role_required

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE")
def company_delete(request, country_slug, company_id):
    country = get_object_or_404(Country, slug=country_slug)
    company = get_object_or_404(Company, pk=company_id)

    if request.method == "POST":
        # 2. TEMPORARILY DISCONNECT THE SIGNAL
        # This tells Django: "Stop checking if the period is completed. Just let me delete it."
        pre_delete.disconnect(prevent_delete_processed_period, sender=PayrollPeriod)
        
        try:
            # 3. Perform the Delete
            company.delete()
            
        finally:
            # 4. RECONNECT THE SIGNAL (Critical!)
            # We must turn the safety system back on for other companies
            pre_delete.connect(prevent_delete_processed_period, sender=PayrollPeriod)

        # 5. Redirect back to the list
        return redirect("company_list", country_slug=country.slug)

    # GET Request: Render the confirmation page
    companies = Company.objects.filter(country=country).order_by("trade_name")

    return render(request, "company/delete.html", {
        "country": country,
        "companies": companies,
        "country_slug": country.slug,
        "company": company,
    })













import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse

# Imports for the enhanced process
from Exactus.country.models import Country
from Exactus.company.models import Company
from Exactus.company.forms import CompanyUploadForm 
from Exactus.country.utils.decorators import role_required

# ────────────────────────────────────────────────────────────────
# 📄 Smart Template Download (Multi-Country Support)
# ────────────────────────────────────────────────────────────────

@login_required
@role_required("EXEC", "ADMIN", "IMPLEMENTATION")
def download_companies_template(request, country_slug=None):
    """
    Generates an Excel template following your specific schema.
    Includes validation dropdowns for Status and Archive choices.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Import"

    # Schema based on your requirements
    headers = [
        "country_code", "company_code", "trade_name", "legal_name", 
        "company_number", "account_status", "account_archive"
    ]
    
    # 1. Apply Headers and Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    # 2. Add Data Validation Dropdowns
    # Account Status
    status_dv = DataValidation(type="list", formula1='"ACTIVE,SUSPENDED,INACTIVE"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add("F2:F500") # Column F

    # Account Archive
    archive_dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(archive_dv)
    archive_dv.add("G2:G500") # Column G

    # 3. Add Example Row
    ws.append(["GB", "EX-001", "Example Ltd", "Example Holdings PLC", "1234567", "ACTIVE", "N"])

    filename = f"Exactus_Bulk_Company_Template.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ────────────────────────────────────────────────────────────────
# 📤 Enhanced Bulk Global Upload (Multi-Country)
# ────────────────────────────────────────────────────────────────

@login_required
@role_required("EXEC", "ADMIN", "IMPLEMENTATION")
@transaction.atomic
def company_upload_view(request, country_slug=None):
    """
    Handles Multi-Country Excel upload.
    Matches country_code in each row to an active Country in the DB.
    """
    if request.method == "POST":
        form = CompanyUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES["file"]
            
            try:
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                
                # Extract headers and validate structure
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    messages.error(request, "The file is empty.")
                    return redirect("companies:company_upload_global")
                
                file_headers = [str(h).strip().lower() for h in rows[0]]
                data_rows = rows[1:]

                success_count = 0
                error_log = []
                
                # Pre-fetch countries for performance
                country_map = {c.iso2_code.upper(): c for c in Country.objects.all()}

                for index, row in enumerate(data_rows, start=2):
                    if not any(row): continue # Skip blank rows
                    
                    # Create a dictionary of the row
                    row_data = dict(zip(file_headers, row))
                    
                    try:
                        # 1. Validate Country Code
                        iso_code = str(row_data.get('country_code', '')).strip().upper()
                        target_country = country_map.get(iso_code)
                        
                        if not target_country:
                            error_log.append(f"Row {index}: Country Code '{iso_code}' not found or inactive.")
                            continue

                        # 2. Prepare Data (Handling Defaults)
                        company_code = str(row_data.get('company_code', '')).strip()
                        trade_name = str(row_data.get('trade_name', '')).strip()
                        legal_name = str(row_data.get('legal_name', '')).strip()
                        
                        if not company_code or not trade_name:
                            error_log.append(f"Row {index}: Missing company_code or trade_name.")
                            continue

                        # 3. Update or Create logic
                        Company.objects.update_or_create(
                            company_code=company_code,
                            country=target_country,
                            defaults={
                                'trade_name': trade_name,
                                'legal_name': legal_name,
                                'company_number': row_data.get('company_number'),
                                'account_status': str(row_data.get('account_status', 'ACTIVE')).upper(),
                                'account_archive': str(row_data.get('account_archive', 'N')).upper(),
                            }
                        )
                        success_count += 1

                    except Exception as e:
                        error_log.append(f"Row {index}: {str(e)}")

                # Final result summary
                request.session["upload_result"] = {
                    "success_count": success_count,
                    "errors": error_log,
                    "total": len(data_rows)
                }
                return redirect("companies:company_upload_result_global")

            except Exception as e:
                messages.error(request, f"Critical error processing Excel: {str(e)}")

    else:
        form = CompanyUploadForm()

    return render(request, "company/upload_form.html", {"form": form})











# Exactus/company/views.py
import csv
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from Exactus.country.models import Country
from Exactus.company.forms import CompanyUploadForm
# (Include your other existing imports here)

@login_required
def company_upload_view(request, country_slug=None):
    country = None
    if country_slug:
        country = get_object_or_404(Country, slug=country_slug)

    is_global = (country is None)

    if request.method == "POST":
        form = CompanyUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = form.cleaned_data["csv_file"]
            dry_run = form.cleaned_data.get("dry_run", False)

            try:
                data_set = csv_file.read().decode("utf-8-sig")
            except UnicodeDecodeError:
                csv_file.seek(0)
                data_set = csv_file.read().decode("iso-8859-1")
            
            import io
            io_string = io.StringIO(data_set)

            # Assuming your importer handles (io_string, country=None) for global
            success_count, error_count, errors = import_companies_from_csv(
                io_string, country=country, dry_run=dry_run
            )

            request.session["upload_results"] = {
                "success_count": success_count,
                "error_count": error_count,
                "errors": errors,
                "country_slug": country_slug,
            }

            if country:
                return redirect(reverse("companies:company_upload_result", kwargs={"country_slug": country_slug}))
            else:
                return redirect("companies:company_upload_result_global")
    else:
        form = CompanyUploadForm()

    return render(
        request,
        "company/upload_form.html",
        {
            "form": form,
            "country": country,
            "is_global": is_global
        }
    )

@login_required
def company_upload_result_view(request, country_slug=None):
    country = None
    if country_slug:
        country = get_object_or_404(Country, slug=country_slug)
        
    results = request.session.pop("upload_results", None)
    if not results:
        if country:
            return redirect("companies:company", country_slug=country_slug)
        return redirect("dashboard")

    return render(request, "company/upload_result.html", {
        "country": country,
        "results": results
    })

@login_required
def download_companies_template(request, country_slug=None):
    """
    Downloads CSV template. 
    - Global Mode (slug=None): Adds 'country_code' column.
    - Local Mode (slug=Set): Pre-fills specific country data.
    """
    country = None
    if country_slug:
        country = get_object_or_404(Country, slug=country_slug)

    response = HttpResponse(content_type='text/csv')
    
    if country:
        filename = f"companies_template_{country.iso2_code}.csv"
    else:
        filename = "companies_template_GLOBAL.csv"
        
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # 1. Define Headers
    headers = []
    if not country:
        headers.append("country_code")
        
    headers.extend([
        "company_code", "company_number", "trade_name", "legal_name",
        "building_name", "road_name_1", "road_name_2", "town", "post_code",
        "tax_id_01", "tax_id_02", "tax_id_03", "tax_id_04", "tax_id_05",
        "rti_user_id", "rti_password", "account_status"
    ])
    
    writer.writerow(headers)

    # 2. Add Example Row
    row = []
    if not country:
        row.append("GB") # Example for Global

    row.extend([
        "COMP001", "12345678", "Example Trading", "Example Ltd",
        "10 Downing St", "Westminster", "", "London", "SW1A 2AA",
        "PAYE123", "", "", "", "",
        "user_rti", "pass123", "ACTIVE"
    ])

    writer.writerow(row)

    return response




def company_test_validation(request, country_slug):
    """Test view to check form validation without saving"""
    country = get_object_or_404(Country, slug=country_slug)
    FormClass = get_company_form_class(country)
    
    if request.method == "POST":
        form = FormClass(request.POST, request.FILES, country=country)
        if form.is_valid():
            return HttpResponse("Form would save successfully!")
        else:
            return HttpResponse(f"Form validation failed. Errors: {form.errors}")
    
    # Show a simple test form
    html = f"""
    <html>
    <body>
        <h1>Test Form Validation for {country.name}</h1>
        <form method="post">
            <input type="hidden" name="country_code" value="{country.iso2_code}">
            <h3>Required Fields:</h3>
    """
    
    test_form = FormClass(country=country)
    
    for field_name, field in test_form.fields.items():
        if field.required:
            html += f"""
            <div style="margin-bottom: 10px;">
                <label>{field.label or field_name} {'(required)' if field.required else ''}</label><br>
                <input type="text" name="{field_name}" placeholder="Enter {field.label or field_name}">
            </div>
            """
    
    html += """
            <button type="submit">Test Validation</button>
        </form>
    </body>
    </html>
    """
    
    return HttpResponse(html)


# ────────────────────────────────────────────────────────────────
# 🐛 DEBUG VIEWS
# ────────────────────────────────────────────────────────────────

def company_debug_info(request, country_slug):
    """Display debug information from previous form submission"""
    country = get_object_or_404(Country, slug=country_slug)
    
    debug_info = request.session.get('debug_info', {})
    validation_errors = request.session.get('validation_errors', {})
    
    # Clear session data after displaying
    if 'debug_info' in request.session:
        del request.session['debug_info']
    if 'validation_errors' in request.session:
        del request.session['validation_errors']
    
    return render(request, "company/debug_info.html", {
        "country": country,
        "debug_info": debug_info,
        "validation_errors": validation_errors,
    })


def company_validate_ajax(request, country_slug):
    """AJAX endpoint for real-time validation"""
    if request.method == "POST" and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        country = get_object_or_404(Country, slug=country_slug)
        FormClass = get_company_form_class(country)
        
        # Create form with POST data
        form = FormClass(request.POST, request.FILES, country=country)
        
        # Check validation
        is_valid = form.is_valid()
        
        # Prepare response
        response_data = {
            'valid': is_valid,
            'errors': form.errors,
            'cleaned_data': form.cleaned_data if is_valid else {},
        }
        
        return JsonResponse(response_data)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


def company_field_requirements(request, country_slug):
    """API endpoint to get field requirements for a country"""
    country = get_object_or_404(Country, slug=country_slug)
    FormClass = get_company_form_class(country)
    
    form = FormClass(country=country)
    
    requirements = {}
    for field_name, field in form.fields.items():
        requirements[field_name] = {
            'label': str(field.label),
            'required': field.required,
            'help_text': str(field.help_text or ''),
            'type': field.__class__.__name__,
        }
    
    return JsonResponse({
        'country': country.iso2_code,
        'requirements': requirements,
    })


# ────────────────────────────────────────────────────────────────
# 📝 FORM TEMPLATE DEBUG VIEW
# ────────────────────────────────────────────────────────────────

def company_form_debug(request, country_slug):
    """Render form with debug information visible"""
    country = get_object_or_404(Country, slug=country_slug)
    FormClass = get_company_form_class(country)
    
    if request.method == "POST":
        form = FormClass(request.POST, request.FILES, country=country)
    else:
        form = FormClass(country=country)
    
    # Analyze form structure
    form_analysis = {
        'total_fields': len(form.fields),
        'required_fields': [],
        'optional_fields': [],
        'hidden_fields': [],
        'choice_fields': [],
    }
    
    for field_name, field in form.fields.items():
        field_info = {
            'name': field_name,
            'label': str(field.label),
            'required': field.required,
            'widget': field.widget.__class__.__name__,
            'help_text': str(field.help_text or ''),
        }
        
        if field.required:
            form_analysis['required_fields'].append(field_info)
        else:
            form_analysis['optional_fields'].append(field_info)
        
        if hasattr(field.widget, 'input_type') and field.widget.input_type == 'hidden':
            form_analysis['hidden_fields'].append(field_info)
        
        if hasattr(field, 'choices') and field.choices:
            form_analysis['choice_fields'].append(field_info)
    
    return render(request, "company/form_debug.html", {
        "form": form,
        "country": country,
        "company": None,
        "form_analysis": form_analysis,
    })


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import ClientGroup
from .forms import ClientGroupForm
from Exactus.utils.decorators import role_required

@login_required
@role_required("EXEC", "ADMIN", "OPERATION")
def client_group_list(request, country_slug):
    """List all client groups."""
    groups = ClientGroup.objects.prefetch_related('companies').all()
    return render(request, 'company/client_group_list.html', {
        'groups': groups,
        'country_slug': country_slug  # Pass to template for breadcrumbs
    })

@login_required
@role_required("EXEC", "ADMIN", "OPERATION")
def client_group_create(request, country_slug):
    """Create a new client group."""
    if request.method == 'POST':
        form = ClientGroupForm(request.POST)
        if form.is_valid():
            group = form.save()
            messages.success(request, f"Client Group '{group.name}' created successfully.")
            return redirect('company:client_group_list', country_slug=country_slug)
    else:
        form = ClientGroupForm()
    
    return render(request, 'company/client_group_form.html', {
        'form': form, 
        'title': 'Create Client Group',
        'country_slug': country_slug
    })

@login_required
@role_required("EXEC", "ADMIN", "OPERATION")
def client_group_edit(request, country_slug, group_id):
    """Edit an existing client group."""
    group = get_object_or_404(ClientGroup, pk=group_id)
    
    if request.method == 'POST':
        form = ClientGroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            messages.success(request, f"Client Group '{group.name}' updated successfully.")
            return redirect('company:client_group_list', country_slug=country_slug)
    else:
        form = ClientGroupForm(instance=group)
    
    return render(request, 'company/client_group_form.html', {
        'form': form, 
        'title': f'Edit {group.name}',
        'country_slug': country_slug
    })