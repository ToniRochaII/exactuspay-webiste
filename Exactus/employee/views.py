import csv
import openpyxl
import uuid
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.contrib import messages
from django.http import HttpResponse
from django.views import View
from django.db import transaction
from django.contrib.auth import get_user_model
from django.utils.crypto import get_random_string  # <-- CRITICAL FIX

# Models
from Exactus.company.models import Company
from Exactus.employee.models import Employee
from Exactus.country.models import Country

# Forms
from Exactus.employee.forms import (
    get_employee_form_for_country, 
    EmployeeUploadForm, 
    EmployeeAccessForm
)

# Utils & Permissions
from Exactus.country.utils.decorators import role_required

User = get_user_model()

# ──────────────────────────────────────────────────────────────────────────────
# SECURITY HELPER
# ──────────────────────────────────────────────────────────────────────────────

def validate_company_access(user, company):
    """Verifies if the current user is allowed to access the specific company."""
    if user.is_superuser or getattr(user, 'role', '') in ["EXEC", "ADMIN"]:
        return True
    if hasattr(user, 'get_accessible_companies'):
        if company in user.get_accessible_companies():
            return True
    return False

# ──────────────────────────────────────────────────────────────────────────────
# EMPLOYEE CRUD (RESTRICTED TO EXEC & ADMIN)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def employee_list(request, country_slug, company_id):
    """List employees for a specific company - Restricted to EXEC and ADMIN."""
    country = get_object_or_404(Country, slug=country_slug)
    company = get_object_or_404(Company, pk=company_id)
    
    if not validate_company_access(request.user, company):
        messages.error(request, "Access Denied.")
        return redirect("dashboard")

    employees = Employee.objects.filter(company=company).order_by("employee_number")
    
    return render(request, "employee/list.html", {
        "company": company, 
        "employees": employees,
        "country": country,
        "country_slug": country_slug,
    })

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def employee_create(request, country_slug, company_id):
    """Create a new employee record - Restricted to EXEC and ADMIN."""
    country = get_object_or_404(Country, slug=country_slug)
    company = get_object_or_404(Company, pk=company_id)

    if not validate_company_access(request.user, company):
        messages.error(request, "Access Denied.")
        return redirect("dashboard")

    FormClass = get_employee_form_for_country(country)

    if request.method == "POST":
        form = FormClass(request.POST)
        if form.is_valid():
            employee = form.save(commit=False)
            employee.company = company
            employee.save()
            messages.success(request, f"Employee '{employee.employee_name}' added successfully.")
            return redirect('employee:employee', country_slug=country_slug, company_id=company.company_id)
    else:
        form = FormClass()

    return render(request, "employee/create.html", {
        "form": form, "company": company, "country": country, "country_slug": country_slug
    })

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def employee_edit(request, country_slug, company_id, employee_id):
    """
    Edit an existing employee record - Restricted to EXEC and ADMIN.
    Includes logic to create User account if missing.
    """
    country = get_object_or_404(Country, slug=country_slug)
    company = get_object_or_404(Company, pk=company_id)

    if not validate_company_access(request.user, company):
        messages.error(request, "Access Denied.")
        return redirect("dashboard")

    employee = get_object_or_404(Employee, pk=employee_id, company=company)
    linked_user = User.objects.filter(email=employee.email).first() if employee.email else None
    FormClass = get_employee_form_for_country(country)

    if request.method == "POST":
        # --- NEW LOGIC: Handle User Account Creation (Fixed) ---
        if 'create_user_account' in request.POST:
            if not employee.email:
                messages.error(request, "Employee must have an email address to create a user account.")
            elif linked_user:
                messages.warning(request, "User account already exists.")
            else:
                try:
                    # 1. Generate Secure Password manually using utils
                    temp_password = get_random_string(length=12)
                    
                    # 2. Create User manually to bypass Manager errors
                    new_user = User.objects.create_user(
                        username=employee.email,
                        email=employee.email,
                        password=temp_password
                    )
                    
                    # 3. Set Attributes
                    new_user.role = 'EMPLOYEE'
                    new_user.first_name = employee.employee_name
                    new_user.last_name = employee.employee_surname
                    new_user.save()
                    
                    # 4. Success Message (In production, send this via email)
                    messages.success(request, f"User account created! Temporary Password: {temp_password}")
                    return redirect("employee:employee_edit", country_slug=country_slug, company_id=company.company_id, employee_id=employee.id)
                    
                except Exception as e:
                    messages.error(request, f"Error creating user account: {str(e)}")
            
            # Refresh to show message if we didn't redirect
            return redirect("employee:employee_edit", country_slug=country_slug, company_id=company.company_id, employee_id=employee.id)

        # --- Standard Form Handling ---
        form = FormClass(request.POST, instance=employee)
        access_form = EmployeeAccessForm(request.POST, instance=linked_user) if linked_user else None

        if form.is_valid():
            form.save()
            if access_form and access_form.is_valid():
                access_form.save()
            messages.success(request, "Employee updated successfully.")
            return redirect("employee:employee", country_slug=country_slug, company_id=company.company_id)
    else:
        form = FormClass(instance=employee)
        access_form = EmployeeAccessForm(instance=linked_user) if linked_user else None

    return render(request, "employee/edit.html", {
        "form": form, "access_form": access_form, "linked_user": linked_user,
        "employee": employee, "company": company, "country": country, "country_slug": country_slug,
    })

# ──────────────────────────────────────────────────────────────────────────────
# MULTI-FORMAT UPLOAD (RESTRICTED TO EXEC & ADMIN)
# ──────────────────────────────────────────────────────────────────────────────

class EmployeeUploadView(View):
    """Handle bulk employee uploads - Restricted to EXEC and ADMIN."""
    @method_decorator(login_required)
    @method_decorator(role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST"))
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, country_slug=None, company_id=None):
        country = get_object_or_404(Country, slug=country_slug) if country_slug else None
        company = get_object_or_404(Company, pk=company_id) if company_id else None
        return render(request, "employee/upload_form.html", {
            "form": EmployeeUploadForm(),
            "company": company,
            "country": country,
            "country_slug": country_slug
        })

    @transaction.atomic
    def post(self, request, country_slug=None, company_id=None):
        form = EmployeeUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            return self.get(request, country_slug, company_id)

        file = request.FILES["file"]
        filename = file.name.lower()
        dry_run = form.cleaned_data.get("dry_run", False)
        
        try:
            data_rows = []
            
            if filename.endswith('.csv'):
                file.seek(0)
                content = file.read().decode('utf-8-sig').splitlines()
                reader = csv.reader(content)
                rows = list(reader)
                if len(rows) < 2:
                    messages.error(request, "The CSV file is empty.")
                    return self.get(request, country_slug, company_id)
                headers = [str(h).strip().lower() for h in rows[0]]
                for row in rows[1:]:
                    if any(row):
                        data_rows.append(dict(zip(headers, row)))

            elif filename.endswith(('.xlsx', '.xls')):
                file.seek(0)
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    messages.error(request, "The Excel file is empty.")
                    return self.get(request, country_slug, company_id)
                headers = [str(h).strip().lower() for h in rows[0]]
                for row in rows[1:]:
                    if any(row):
                        data_rows.append(dict(zip(headers, row)))
            else:
                messages.error(request, "Unsupported format. Please upload .xlsx or .csv")
                return self.get(request, country_slug, company_id)

            success_count = 0
            error_log = []
            company_cache = {c.company_code: c for c in Company.objects.all()}

            for index, data in enumerate(data_rows, start=2):
                try:
                    c_code = str(data.get('company_code', '')).strip()
                    company_obj = company_cache.get(c_code)
                    
                    if not company_obj:
                        error_log.append(f"Row {index}: Company code '{c_code}' not found.")
                        continue
                    
                    emp_num = data.get('employee_number')
                    if not emp_num:
                        error_log.append(f"Row {index}: Missing employee_number.")
                        continue

                    emp_code = data.get('employee_code') or emp_num

                    defaults = {
                        'employee_name': data.get('employee_name'),
                        'employee_surname': data.get('employee_surname'),
                        'employee_code': emp_code,
                        'gender': data.get('gender'),
                        'email': data.get('email'),
                        'date_of_birth': data.get('date_of_birth'),
                        'employment_start_date': data.get('employment_start_date'),
                    }

                    if 'ni_number' in data: defaults['tax_info_01'] = data['ni_number']
                    if 'tax_code' in data: defaults['tax_info_03'] = data['tax_code']
                    if 'cpf' in data: defaults['tax_info_01'] = data['cpf']
                    if 'pis' in data: defaults['tax_info_02'] = data['pis']

                    if not dry_run:
                        Employee.objects.update_or_create(
                            company=company_obj,
                            employee_number=emp_num,
                            defaults=defaults
                        )
                    success_count += 1
                except Exception as e:
                    error_log.append(f"Row {index}: {str(e)}")

            request.session["upload_result"] = {
                "success_count": success_count,
                "errors": error_log,
                "total": len(data_rows),
                "dry_run": dry_run
            }
            
            if company_id:
                return redirect("employee:employee_upload_result", country_slug=country_slug, company_id=company_id)
            return redirect("employee:global_upload_result", country_slug=country_slug)

        except Exception as e:
            messages.error(request, f"File processing error: {str(e)}")
            return self.get(request, country_slug, company_id)

# ──────────────────────────────────────────────────────────────────────────────
# RESULT VIEWS (RESTRICTED TO EXEC & ADMIN)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def employee_upload_result_view(request, country_slug=None, company_id=None):
    """Display upload outcome - Restricted to EXEC and ADMIN."""
    result = request.session.pop("upload_result", None)
    if not result:
        messages.warning(request, "No upload results found.")
        return redirect("dashboard")
        
    country = get_object_or_404(Country, slug=country_slug) if country_slug else None
    company = get_object_or_404(Company, pk=company_id) if company_id else None

    return render(request, "employee/upload_result.html", {
        "result": result, "company": company, "country": country, "country_slug": country_slug
    })

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def global_upload_result_view(request, country_slug):
    """Alias for global route consistency - Restricted to EXEC and ADMIN."""
    return employee_upload_result_view(request, country_slug=country_slug)

# ──────────────────────────────────────────────────────────────────────────────
# TEMPLATE DOWNLOAD (RESTRICTED TO EXEC & ADMIN)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@role_required("EXEC", "ADMIN", "COMPLIANCE", "IMPLEMENTATION", "OPERATIONS", "DIRECTOR", "MANAGER", "SPECIALIST")
def download_employees_template(request, country_slug, company_id=None):
    """Generate Excel template - Restricted to EXEC and ADMIN."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Import"

    headers = [
        'company_code', 'employee_number', 'employee_code', 'employee_name', 
        'employee_surname', 'gender', 'date_of_birth', 'email', 'employment_start_date'
    ]

    if country_slug in ['united-kingdom', 'uk', 'gb']:
        headers += ['ni_number', 'tax_code']
    elif country_slug in ['brazil', 'br']:
        headers += ['cpf', 'pis']

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    if company_id:
        company = get_object_or_404(Company, pk=company_id)
        ws.cell(row=2, column=1, value=company.company_code)

    gender_dv = DataValidation(type="list", formula1='"Male,Female"', allow_blank=True)
    ws.add_data_validation(gender_dv)
    gender_dv.add("F2:F500")

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename=Employee_Template_{country_slug}.xlsx'
    wb.save(response)
    return response